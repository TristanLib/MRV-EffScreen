#!/usr/bin/env python3
"""Build the unified THETIS-MRV modeling dataset.

Outputs:
- data/interim/mrv_unified_public_reports.csv
- data/processed/mrv_modeling_base.csv
- reports/tables/mrv_processed_missingness.csv
- reports/tables/mrv_label_distribution.csv
- reports/tables/mrv_label_coverage_by_year.csv
- reports/tables/mrv_label_group_coverage.csv
- reports/tables/mrv_year_scope_counts.csv
- reports/tables/mrv_processed_summary.csv
- reports/figures/*.svg
"""

from __future__ import annotations

import csv
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from statistics import median
from typing import Any, Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
RAW_DIR = ROOT / "data" / "raw" / "mrv"
INTERIM_DIR = ROOT / "data" / "interim"
PROCESSED_DIR = ROOT / "data" / "processed"
TABLE_DIR = ROOT / "reports" / "tables"
FIGURE_DIR = ROOT / "reports" / "figures"
LABEL_GROUP_MIN_ROWS = 30


PROCESSED_FIELDS = [
    "source_file",
    "sheet",
    "report_scope",
    "reporting_year",
    "reporting_period_raw",
    "is_full_year",
    "is_main_experiment",
    "temporal_split",
    "imo_number",
    "ship_name",
    "ship_type",
    "technical_efficiency_raw",
    "technical_efficiency_type",
    "technical_efficiency_value",
    "technical_efficiency_is_not_applicable",
    "port_of_registry",
    "has_home_port",
    "has_ice_class",
    "monitoring_method_a",
    "monitoring_method_b",
    "monitoring_method_c",
    "monitoring_method_d_1",
    "monitoring_method_d_2",
    "total_fuel_consumption_mt",
    "fuel_consumption_on_laden_mt",
    "total_co2_emissions_mt",
    "co2_between_ms_ports_mt",
    "co2_departed_ms_ports_mt",
    "co2_to_ms_ports_mt",
    "co2_at_berth_mt",
    "co2_within_ports_mt",
    "co2_on_laden_mt",
    "co2_passenger_transport_mt",
    "co2_freight_transport_mt",
    "time_spent_at_sea_hours",
    "distance_through_ice_nm",
    "time_spent_at_sea_through_ice_hours",
    "fuel_per_distance_kg_nm",
    "fuel_per_distance_on_laden_kg_nm",
    "fuel_per_transport_work_mass_g_tnm",
    "fuel_per_transport_work_volume_g_m3nm",
    "fuel_per_transport_work_dwt_g_dwtnm",
    "fuel_per_transport_work_pax_g_paxnm",
    "fuel_per_transport_work_freight_g_tnm",
    "co2_per_distance_kg_nm",
    "co2_per_distance_on_laden_kg_nm",
    "co2_per_transport_work_mass_g_tnm",
    "co2_per_transport_work_volume_g_m3nm",
    "co2_per_transport_work_dwt_g_dwtnm",
    "co2_per_transport_work_pax_g_paxnm",
    "co2_per_transport_work_freight_g_tnm",
    "average_density_cargo_mt_m3",
    "distance_efficiency_group_n",
    "distance_efficiency_rank_pct",
    "efficiency_label_distance",
    "efficiency_label_distance_code",
]


ALIASES = {
    "imo_number": ["ship__imo_number"],
    "ship_name": ["ship__name"],
    "ship_type": ["ship__ship_type"],
    "reporting_period_raw": ["ship__reporting_period"],
    "technical_efficiency_raw": ["ship__technical_efficiency"],
    "port_of_registry": ["ship__port_of_registry"],
    "home_port": ["ship__home_port"],
    "ice_class": ["ship__ice_class"],
    "monitoring_method_a": ["monitoring_methods__a"],
    "monitoring_method_b": ["monitoring_methods__b"],
    "monitoring_method_c": ["monitoring_methods__c"],
    "monitoring_method_d_1": ["monitoring_methods__d"],
    "monitoring_method_d_2": ["monitoring_methods__d_2"],
    "total_fuel_consumption_mt": [
        "annual_monitoring_results__total_fuel_consumption",
    ],
    "fuel_consumption_on_laden_mt": [
        "annual_monitoring_results__fuel_consumptions_assigned_to_on_laden",
    ],
    "total_co2_emissions_mt": ["annual_monitoring_results__total_co2_emissions"],
    "co2_between_ms_ports_mt": [
        "annual_monitoring_results__co2_emissions_from_all_voyages_between_ports_under_a_ms_jurisdiction",
    ],
    "co2_departed_ms_ports_mt": [
        "annual_monitoring_results__co2_emissions_from_all_voyages_which_departed_from_ports_under_a_ms_jurisdiction",
    ],
    "co2_to_ms_ports_mt": [
        "annual_monitoring_results__co2_emissions_from_all_voyages_to_ports_under_a_ms_jurisdiction",
    ],
    "co2_at_berth_mt": [
        "annual_monitoring_results__co2_emissions_which_occurred_within_ports_under_a_ms_jurisdiction_at_berth",
    ],
    "co2_within_ports_mt": [
        "annual_monitoring_results__co2_emissions_which_occurred_within_ports_under_a_ms_jurisdiction",
    ],
    "co2_on_laden_mt": [
        "annual_monitoring_results__co2_emissions_assigned_to_on_laden",
    ],
    "co2_passenger_transport_mt": [
        "annual_monitoring_results__co2_emissions_assigned_to_passenger_transport",
    ],
    "co2_freight_transport_mt": [
        "annual_monitoring_results__co2_emissions_assigned_to_freight_transport",
    ],
    "time_spent_at_sea_hours": [
        "annual_monitoring_results__annual_total_time_spent_at_sea",
        "annual_monitoring_results__annual_time_spent_at_sea",
        "annual_monitoring_results__time_spent_at_sea",
    ],
    "distance_through_ice_nm": [
        "voluntary_reporting__through_ice",
        "annual_monitoring_results__distance_through_ice",
    ],
    "time_spent_at_sea_through_ice_hours": [
        "voluntary_reporting__total_time_spent_at_sea_through_ice",
        "annual_monitoring_results__time_spent_at_sea_through_ice",
    ],
    "fuel_per_distance_kg_nm": [
        "annual_monitoring_results__annual_average_fuel_consumption_per_distance",
        "annual_monitoring_results__fuel_consumption_per_distance",
    ],
    "fuel_per_distance_on_laden_kg_nm": [
        "voluntary_reporting__fuel_consumption_per_distance_on_laden_voyages",
        "annual_monitoring_results__fuel_consumption_per_distance_on_laden_voyages",
    ],
    "fuel_per_transport_work_mass_g_tnm": [
        "annual_monitoring_results__annual_average_fuel_consumption_per_transport_work_mass",
        "annual_monitoring_results__fuel_consumption_per_transport_work_mass",
    ],
    "fuel_per_transport_work_volume_g_m3nm": [
        "annual_monitoring_results__annual_average_fuel_consumption_per_transport_work_volume",
        "annual_monitoring_results__fuel_consumption_per_transport_work_volume",
    ],
    "fuel_per_transport_work_dwt_g_dwtnm": [
        "annual_monitoring_results__annual_average_fuel_consumption_per_transport_work_dwt",
        "annual_monitoring_results__fuel_consumption_per_transport_work_dwt",
    ],
    "fuel_per_transport_work_pax_g_paxnm": [
        "annual_monitoring_results__annual_average_fuel_consumption_per_transport_work_pax",
        "annual_monitoring_results__fuel_consumption_per_transport_work_pax",
    ],
    "fuel_per_transport_work_freight_g_tnm": [
        "annual_monitoring_results__annual_average_fuel_consumption_per_transport_work_freight",
        "annual_monitoring_results__fuel_consumption_per_transport_work_freight",
    ],
    "co2_per_distance_kg_nm": [
        "annual_monitoring_results__annual_average_co2_emissions_per_distance",
        "annual_monitoring_results__co2_emissions_per_distance",
    ],
    "co2_per_distance_on_laden_kg_nm": [
        "voluntary_reporting__co2_emissions_per_distance_on_laden_voyages",
        "annual_monitoring_results__co2_emissions_per_distance_on_laden_voyages",
    ],
    "co2_per_transport_work_mass_g_tnm": [
        "annual_monitoring_results__annual_average_co2_emissions_per_transport_work_mass",
        "annual_monitoring_results__co2_emissions_per_transport_work_mass",
    ],
    "co2_per_transport_work_volume_g_m3nm": [
        "annual_monitoring_results__annual_average_co2_emissions_per_transport_work_volume",
        "annual_monitoring_results__co2_emissions_per_transport_work_volume",
    ],
    "co2_per_transport_work_dwt_g_dwtnm": [
        "annual_monitoring_results__annual_average_co2_emissions_per_transport_work_dwt",
        "annual_monitoring_results__co2_emissions_per_transport_work_dwt",
    ],
    "co2_per_transport_work_pax_g_paxnm": [
        "annual_monitoring_results__annual_average_co2_emissions_per_transport_work_pax",
        "annual_monitoring_results__co2_emissions_per_transport_work_pax",
    ],
    "co2_per_transport_work_freight_g_tnm": [
        "annual_monitoring_results__annual_average_co2_emissions_per_transport_work_freight",
        "annual_monitoring_results__co2_emissions_per_transport_work_freight",
    ],
    "average_density_cargo_mt_m3": [
        "voluntary_reporting__average_density_of_the_cargo_transported",
        "annual_monitoring_results__average_density_of_the_cargo_transported",
    ],
}


NUMERIC_FIELDS = {
    "technical_efficiency_value",
    "total_fuel_consumption_mt",
    "fuel_consumption_on_laden_mt",
    "total_co2_emissions_mt",
    "co2_between_ms_ports_mt",
    "co2_departed_ms_ports_mt",
    "co2_to_ms_ports_mt",
    "co2_at_berth_mt",
    "co2_within_ports_mt",
    "co2_on_laden_mt",
    "co2_passenger_transport_mt",
    "co2_freight_transport_mt",
    "time_spent_at_sea_hours",
    "distance_through_ice_nm",
    "time_spent_at_sea_through_ice_hours",
    "fuel_per_distance_kg_nm",
    "fuel_per_distance_on_laden_kg_nm",
    "fuel_per_transport_work_mass_g_tnm",
    "fuel_per_transport_work_volume_g_m3nm",
    "fuel_per_transport_work_dwt_g_dwtnm",
    "fuel_per_transport_work_pax_g_paxnm",
    "fuel_per_transport_work_freight_g_tnm",
    "co2_per_distance_kg_nm",
    "co2_per_distance_on_laden_kg_nm",
    "co2_per_transport_work_mass_g_tnm",
    "co2_per_transport_work_volume_g_m3nm",
    "co2_per_transport_work_dwt_g_dwtnm",
    "co2_per_transport_work_pax_g_paxnm",
    "co2_per_transport_work_freight_g_tnm",
    "average_density_cargo_mt_m3",
    "distance_efficiency_group_n",
    "distance_efficiency_rank_pct",
    "efficiency_label_distance_code",
}


@dataclass
class FigureDatum:
    label: str
    value: float


def normalize(value: str) -> str:
    value = value.strip().lower()
    replacements = {
        "co₂eq": "co2eq",
        "co₂": "co2",
        "ch₄": "ch4",
        "n₂o": "n2o",
        "·": " ",
        "³": "3",
    }
    for old, new in replacements.items():
        value = value.replace(old, new)
    value = re.sub(r"\[[^\]]+\]", "", value)
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return value.strip("_")


def stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\n", " ").strip()


def parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return float(value)
    text = stringify(value)
    if not text or text.lower() in {"not applicable", "division by zero!", "nan", "n/a"}:
        return None
    text = text.replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    return float(match.group(0))


def parse_reporting_year(value: Any, fallback: int | None) -> int | None:
    if isinstance(value, (int, float)):
        return int(value)
    match = re.search(r"\b(20\d{2})\b", stringify(value))
    if match:
        return int(match.group(1))
    return fallback


def parse_technical_efficiency(value: Any) -> tuple[str, float | None, str]:
    text = stringify(value)
    if not text:
        return "", None, "false"
    if text.lower() == "not applicable":
        return "NOT_APPLICABLE", None, "true"
    match = re.match(r"\s*([A-Za-z0-9]+)\s*(?:\((.*?)\))?\s*$", text)
    if not match:
        return "", parse_float(text), "false"
    return match.group(1).upper(), parse_float(match.group(2) or ""), "false"


def sheet_scope(sheet_name: str) -> str:
    lowered = sheet_name.lower()
    if "partial" in lowered:
        return "partial_er"
    if "full" in lowered:
        return "full_er"
    return "annual_er"


def temporal_split(year: int | None, scope: str) -> str:
    if scope == "partial_er":
        return "excluded_partial_2024"
    if scope == "full_er" and year == 2024:
        return "external_2024"
    if scope != "annual_er":
        return "excluded"
    if year in {2018, 2019, 2020, 2021}:
        return "train"
    if year == 2022:
        return "holdout_2022"
    if year == 2023:
        return "test_2023"
    return "excluded"


def flattened_headers(ws) -> list[tuple[int, str]]:
    group_row = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    field_row = list(next(ws.iter_rows(min_row=3, max_row=3, values_only=True)))
    group = ""
    seen = Counter()
    headers = []
    for idx, (raw_group, raw_field) in enumerate(zip(group_row, field_row), start=1):
        if raw_group is not None:
            group = stringify(raw_group)
        if raw_field is None:
            continue
        normalized = f"{normalize(group)}__{normalize(stringify(raw_field))}"
        seen[normalized] += 1
        if seen[normalized] > 1:
            normalized = f"{normalized}_{seen[normalized]}"
        headers.append((idx - 1, normalized))
    return headers


def get_first(row_map: dict[str, Any], aliases: list[str]) -> Any:
    for alias in aliases:
        value = row_map.get(alias)
        if value is not None and value != "":
            return value
    return None


def read_workbook_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for path in sorted(RAW_DIR.glob("*.xlsx")):
        fallback_year = int(path.name[:4]) if re.match(r"\d{4}", path.name) else None
        workbook = load_workbook(path, read_only=True, data_only=True)
        for ws in workbook.worksheets:
            scope = sheet_scope(ws.title)
            headers = flattened_headers(ws)
            for raw_row in ws.iter_rows(min_row=4, values_only=True):
                if not any(value is not None and value != "" for value in raw_row):
                    continue
                row_map = {
                    normalized: raw_row[idx] if idx < len(raw_row) else None
                    for idx, normalized in headers
                }
                reporting_period_raw = stringify(get_first(row_map, ALIASES["reporting_period_raw"]))
                reporting_year = parse_reporting_year(reporting_period_raw, fallback_year)
                tech_type, tech_value, tech_na = parse_technical_efficiency(
                    get_first(row_map, ALIASES["technical_efficiency_raw"])
                )
                out: dict[str, str] = {
                    "source_file": path.name,
                    "sheet": ws.title,
                    "report_scope": scope,
                    "reporting_year": stringify(reporting_year),
                    "reporting_period_raw": reporting_period_raw,
                    "is_full_year": "true" if scope != "partial_er" else "false",
                    "is_main_experiment": "true"
                    if scope == "annual_er" and reporting_year in {2018, 2019, 2020, 2021, 2022, 2023}
                    else "false",
                    "temporal_split": temporal_split(reporting_year, scope),
                    "technical_efficiency_type": tech_type,
                    "technical_efficiency_value": format_number(tech_value),
                    "technical_efficiency_is_not_applicable": tech_na,
                }
                for field in [
                    "imo_number",
                    "ship_name",
                    "ship_type",
                    "technical_efficiency_raw",
                    "port_of_registry",
                    "monitoring_method_a",
                    "monitoring_method_b",
                    "monitoring_method_c",
                    "monitoring_method_d_1",
                    "monitoring_method_d_2",
                ]:
                    out[field] = stringify(get_first(row_map, ALIASES[field]))
                out["has_home_port"] = "true" if stringify(get_first(row_map, ALIASES["home_port"])) else "false"
                out["has_ice_class"] = "true" if stringify(get_first(row_map, ALIASES["ice_class"])) else "false"
                for field in sorted(set(ALIASES) & NUMERIC_FIELDS):
                    out[field] = format_number(parse_float(get_first(row_map, ALIASES[field])))
                rows.append({field: out.get(field, "") for field in PROCESSED_FIELDS})
    return rows


def add_distance_efficiency_labels(rows: list[dict[str, str]]) -> None:
    groups: dict[tuple[str, str], list[tuple[int, float]]] = defaultdict(list)
    for idx, row in enumerate(rows):
        value = parse_float(row.get("co2_per_distance_kg_nm"))
        if value is None:
            continue
        key = (row.get("ship_type", ""), row.get("reporting_year", ""))
        if not key[0] or not key[1]:
            continue
        groups[key].append((idx, value))

    for group_rows in groups.values():
        group_rows.sort(key=lambda item: item[1])
        n = len(group_rows)
        if n < LABEL_GROUP_MIN_ROWS:
            continue
        for rank, (idx, _value) in enumerate(group_rows):
            pct = (rank + 1) / n
            if rank < n / 3:
                label, code = "efficient", "0"
            elif rank < 2 * n / 3:
                label, code = "medium", "1"
            else:
                label, code = "inefficient", "2"
            rows[idx]["distance_efficiency_group_n"] = str(n)
            rows[idx]["distance_efficiency_rank_pct"] = f"{pct:.6f}"
            rows[idx]["efficiency_label_distance"] = label
            rows[idx]["efficiency_label_distance_code"] = code


def format_number(value: float | None) -> str:
    if value is None:
        return ""
    return f"{value:.10g}"


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = fields or (list(rows[0]) if rows else [])
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def build_missingness(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    out = []
    total = len(rows)
    for field in PROCESSED_FIELDS:
        missing = sum(1 for row in rows if row.get(field, "") == "")
        role = field_role(field)
        out.append(
            {
                "field": field,
                "role": role,
                "non_missing": str(total - missing),
                "missing": str(missing),
                "missing_pct": f"{missing / total:.4f}" if total else "1.0000",
                "recommended_handling": recommended_handling(field, missing / total if total else 1),
            }
        )
    return out


def field_role(field: str) -> str:
    if field in {"efficiency_label_distance", "efficiency_label_distance_code"}:
        return "target"
    if field in {"source_file", "sheet", "report_scope", "temporal_split", "is_main_experiment"}:
        return "metadata"
    if field.startswith("co2_") or field.startswith("fuel_") or field.startswith("total_"):
        return "leakage_or_screening"
    if field.startswith("technical_efficiency") or field.startswith("monitoring_method"):
        return "candidate_feature"
    return "candidate_feature"


def recommended_handling(field: str, missing_pct: float) -> str:
    if field in {"ship_name", "imo_number", "source_file", "sheet"}:
        return "retain_for_traceability_not_model_feature"
    if field in {"efficiency_label_distance", "efficiency_label_distance_code"}:
        return "drop_unlabeled_rows_for_classification"
    if field in {"has_home_port", "has_ice_class", "technical_efficiency_is_not_applicable"}:
        return "use_as_boolean"
    if field.startswith("co2_") or field.startswith("fuel_") or field.startswith("total_"):
        return "exclude_from_strict_classification_features; allow_for_screening"
    if missing_pct >= 0.7:
        return "drop_or_binarize"
    if missing_pct >= 0.2:
        return "impute_with_missing_indicator"
    if field in NUMERIC_FIELDS:
        return "median_impute_within_train_or_ship_type"
    return "mode_or_unknown_category"


def build_label_distribution(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    counter = Counter()
    for row in rows:
        label = row.get("efficiency_label_distance") or "unlabeled"
        counter[(row.get("reporting_year", ""), row.get("report_scope", ""), label)] += 1
    return [
        {
            "reporting_year": year,
            "report_scope": scope,
            "efficiency_label_distance": label,
            "rows": str(count),
        }
        for (year, scope, label), count in sorted(counter.items())
    ]


def build_label_group_coverage(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    groups: dict[tuple[str, str, str, str], dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for row in rows:
        year = row.get("reporting_year", "")
        ship_type = row.get("ship_type", "")
        if not year or not ship_type:
            continue
        key = (year, row.get("report_scope", ""), row.get("temporal_split", ""), ship_type)
        groups[key]["rows"] += 1
        if parse_float(row.get("co2_per_distance_kg_nm")) is not None:
            groups[key]["target_non_missing_rows"] += 1
        if row.get("efficiency_label_distance"):
            groups[key]["labeled_rows"] += 1

    out = []
    for (year, scope, split, ship_type), counts in sorted(groups.items()):
        rows_total = counts["rows"]
        target_rows = counts["target_non_missing_rows"]
        labeled_rows = counts["labeled_rows"]
        missing_target_rows = rows_total - target_rows
        below_threshold_rows = target_rows if 0 < target_rows < LABEL_GROUP_MIN_ROWS else 0
        if target_rows >= LABEL_GROUP_MIN_ROWS and labeled_rows == target_rows:
            status = "labeled"
        elif target_rows >= LABEL_GROUP_MIN_ROWS:
            status = "partially_unlabeled_missing_target"
        elif target_rows > 0:
            status = "below_min_target_rows"
        else:
            status = "no_target_values"
        out.append(
            {
                "reporting_year": year,
                "report_scope": scope,
                "temporal_split": split,
                "ship_type": ship_type,
                "rows": str(rows_total),
                "target_non_missing_rows": str(target_rows),
                "labeled_rows": str(labeled_rows),
                "unlabeled_rows": str(rows_total - labeled_rows),
                "missing_target_rows": str(missing_target_rows),
                "below_threshold_target_rows": str(below_threshold_rows),
                "labeling_status": status,
            }
        )
    return out


def build_label_coverage_by_year(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    counter: dict[tuple[str, str, str], dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for row in rows:
        key = (row.get("reporting_year", ""), row.get("report_scope", ""), row.get("temporal_split", ""))
        counter[key]["rows"] += 1
        if row.get("efficiency_label_distance"):
            counter[key]["labeled_rows"] += 1
    out = []
    for (year, scope, split), counts in sorted(counter.items()):
        total = counts["rows"]
        labeled = counts["labeled_rows"]
        out.append(
            {
                "reporting_year": year,
                "report_scope": scope,
                "temporal_split": split,
                "rows": str(total),
                "labeled_rows": str(labeled),
                "unlabeled_rows": str(total - labeled),
                "label_coverage_pct": f"{labeled / total:.6f}" if total else "0.000000",
            }
        )
    return out


def build_year_scope_counts(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    counter = Counter((row["reporting_year"], row["report_scope"], row["temporal_split"]) for row in rows)
    return [
        {
            "reporting_year": year,
            "report_scope": scope,
            "temporal_split": split,
            "rows": str(count),
        }
        for (year, scope, split), count in sorted(counter.items())
    ]


def build_summary(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    labeled_main = [
        row
        for row in rows
        if row["is_main_experiment"] == "true" and row["efficiency_label_distance"]
    ]
    main_rows = [row for row in rows if row["is_main_experiment"] == "true"]
    main_unlabeled = [row for row in main_rows if not row["efficiency_label_distance"]]
    group_coverage = build_label_group_coverage(rows)
    main_groups = [
        row
        for row in group_coverage
        if row["report_scope"] == "annual_er"
        and row["reporting_year"] in {"2018", "2019", "2020", "2021", "2022", "2023"}
    ]
    below_threshold_groups = [row for row in main_groups if row["labeling_status"] == "below_min_target_rows"]
    values = [parse_float(row["co2_per_distance_kg_nm"]) for row in labeled_main]
    values = [value for value in values if value is not None]
    return [
        {"metric": "total_rows", "value": str(len(rows))},
        {"metric": "main_experiment_rows", "value": str(len(main_rows))},
        {"metric": "labeled_main_experiment_rows", "value": str(len(labeled_main))},
        {"metric": "unlabeled_main_experiment_rows", "value": str(len(main_unlabeled))},
        {"metric": "labeling_min_ship_type_year_rows", "value": str(LABEL_GROUP_MIN_ROWS)},
        {"metric": "main_ship_type_year_groups", "value": str(len(main_groups))},
        {"metric": "main_below_threshold_groups", "value": str(len(below_threshold_groups))},
        {
            "metric": "main_below_threshold_target_rows",
            "value": str(sum(int(row["below_threshold_target_rows"]) for row in below_threshold_groups)),
        },
        {
            "metric": "main_missing_target_rows",
            "value": str(sum(int(row["missing_target_rows"]) for row in main_groups)),
        },
        {
            "metric": "external_2024_rows",
            "value": str(sum(1 for row in rows if row["temporal_split"] == "external_2024")),
        },
        {
            "metric": "partial_er_rows",
            "value": str(sum(1 for row in rows if row["report_scope"] == "partial_er")),
        },
        {"metric": "unique_imo_numbers", "value": str(len({row["imo_number"] for row in rows if row["imo_number"]}))},
        {"metric": "unique_ship_types", "value": str(len({row["ship_type"] for row in rows if row["ship_type"]}))},
        {
            "metric": "median_co2_per_distance_labeled_main",
            "value": format_number(median(values) if values else None),
        },
    ]


def make_figures(rows: list[dict[str, str]]) -> None:
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)
    year_counts = Counter(row["reporting_year"] for row in rows if row["reporting_year"])
    draw_bar_chart(
        FIGURE_DIR / "mrv_rows_by_year.svg",
        [FigureDatum(year, count) for year, count in sorted(year_counts.items())],
        "Rows by reporting year",
        "Rows",
    )

    ship_counts = Counter(row["ship_type"] for row in rows if row["ship_type"])
    draw_bar_chart(
        FIGURE_DIR / "mrv_top_ship_types.svg",
        [FigureDatum(label, count) for label, count in ship_counts.most_common(10)],
        "Top ship types",
        "Rows",
        width=920,
        height=520,
    )

    label_counts = Counter(
        (row["reporting_year"], row["efficiency_label_distance"])
        for row in rows
        if row["is_main_experiment"] == "true" and row["efficiency_label_distance"]
    )
    label_data = []
    for year in sorted({year for year, _ in label_counts}):
        for label in ["efficient", "medium", "inefficient"]:
            label_data.append(FigureDatum(f"{year} {label}", label_counts[(year, label)]))
    draw_bar_chart(
        FIGURE_DIR / "mrv_label_distribution_by_year.svg",
        label_data,
        "Distance-efficiency labels by year",
        "Rows",
        width=1100,
        height=560,
        rotate_labels=True,
    )

    coverage_data = []
    for row in build_label_coverage_by_year(rows):
        if row["report_scope"] == "annual_er" and row["reporting_year"] in {"2018", "2019", "2020", "2021", "2022", "2023"}:
            coverage_data.append(FigureDatum(row["reporting_year"], float(row["label_coverage_pct"])))
    draw_bar_chart(
        FIGURE_DIR / "mrv_label_coverage_by_year.svg",
        coverage_data,
        "Label coverage by reporting year",
        "coverage fraction",
        width=860,
        height=460,
    )

    medians = []
    for ship_type, _count in ship_counts.most_common(10):
        values = [
            parse_float(row["co2_per_distance_kg_nm"])
            for row in rows
            if row["ship_type"] == ship_type and row["is_main_experiment"] == "true"
        ]
        values = [value for value in values if value is not None]
        if values:
            medians.append(FigureDatum(ship_type, median(values)))
    draw_bar_chart(
        FIGURE_DIR / "mrv_median_co2_per_distance_top_ship_types.svg",
        medians,
        "Median CO2 per distance, top ship types",
        "kg CO2 / n mile",
        width=920,
        height=520,
    )


def draw_bar_chart(
    path: Path,
    data: list[FigureDatum],
    title: str,
    y_label: str,
    width: int = 860,
    height: int = 460,
    rotate_labels: bool = False,
) -> None:
    margin_left, margin_right, margin_top, margin_bottom = 78, 26, 54, 110
    plot_w = width - margin_left - margin_right
    plot_h = height - margin_top - margin_bottom
    max_value = max((datum.value for datum in data), default=1)
    bar_gap = 6
    bar_w = max(4, (plot_w - bar_gap * max(0, len(data) - 1)) / max(1, len(data)))
    colors = ["#2f6f73", "#8f5f2a", "#4f6f9f", "#a64d4d", "#5e6c84"]

    parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" width="{width}" height="{height}" viewBox="0 0 {width} {height}">',
        '<rect width="100%" height="100%" fill="#ffffff"/>',
        f'<text x="{width / 2}" y="28" text-anchor="middle" font-family="Arial" font-size="18" fill="#1f2933">{escape_xml(title)}</text>',
        f'<text x="18" y="{margin_top + plot_h / 2}" transform="rotate(-90 18 {margin_top + plot_h / 2})" text-anchor="middle" font-family="Arial" font-size="12" fill="#52606d">{escape_xml(y_label)}</text>',
        f'<line x1="{margin_left}" y1="{margin_top + plot_h}" x2="{width - margin_right}" y2="{margin_top + plot_h}" stroke="#9aa5b1" stroke-width="1"/>',
        f'<line x1="{margin_left}" y1="{margin_top}" x2="{margin_left}" y2="{margin_top + plot_h}" stroke="#9aa5b1" stroke-width="1"/>',
    ]

    for tick in range(5):
        value = max_value * tick / 4
        y = margin_top + plot_h - (value / max_value) * plot_h if max_value else margin_top + plot_h
        parts.append(f'<line x1="{margin_left - 4}" y1="{y:.2f}" x2="{width - margin_right}" y2="{y:.2f}" stroke="#e4e7eb" stroke-width="1"/>')
        parts.append(f'<text x="{margin_left - 8}" y="{y + 4:.2f}" text-anchor="end" font-family="Arial" font-size="11" fill="#52606d">{format_axis(value)}</text>')

    for i, datum in enumerate(data):
        x = margin_left + i * (bar_w + bar_gap)
        bar_h = (datum.value / max_value) * plot_h if max_value else 0
        y = margin_top + plot_h - bar_h
        color = colors[i % len(colors)]
        parts.append(f'<rect x="{x:.2f}" y="{y:.2f}" width="{bar_w:.2f}" height="{bar_h:.2f}" fill="{color}"/>')
        label_x = x + bar_w / 2
        label_y = margin_top + plot_h + 18
        label = truncate_label(datum.label, 24 if rotate_labels else 18)
        if rotate_labels:
            parts.append(f'<text x="{label_x:.2f}" y="{label_y:.2f}" transform="rotate(55 {label_x:.2f} {label_y:.2f})" text-anchor="start" font-family="Arial" font-size="10" fill="#323f4b">{escape_xml(label)}</text>')
        else:
            parts.append(f'<text x="{label_x:.2f}" y="{label_y:.2f}" text-anchor="middle" font-family="Arial" font-size="10" fill="#323f4b">{escape_xml(label)}</text>')
    parts.append("</svg>")
    path.write_text("\n".join(parts), encoding="utf-8")


def format_axis(value: float) -> str:
    if value >= 1000:
        return f"{value / 1000:.0f}k"
    if value >= 10:
        return f"{value:.0f}"
    return f"{value:.1f}"


def truncate_label(value: str, length: int) -> str:
    return value if len(value) <= length else value[: length - 1] + "…"


def escape_xml(value: str) -> str:
    return (
        value.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def main() -> None:
    for directory in [INTERIM_DIR, PROCESSED_DIR, TABLE_DIR, FIGURE_DIR]:
        directory.mkdir(parents=True, exist_ok=True)

    rows = read_workbook_rows()
    add_distance_efficiency_labels(rows)

    write_csv(INTERIM_DIR / "mrv_unified_public_reports.csv", rows, PROCESSED_FIELDS)
    write_csv(PROCESSED_DIR / "mrv_modeling_base.csv", rows, PROCESSED_FIELDS)
    write_csv(TABLE_DIR / "mrv_processed_missingness.csv", build_missingness(rows))
    write_csv(TABLE_DIR / "mrv_label_distribution.csv", build_label_distribution(rows))
    write_csv(TABLE_DIR / "mrv_label_coverage_by_year.csv", build_label_coverage_by_year(rows))
    write_csv(TABLE_DIR / "mrv_label_group_coverage.csv", build_label_group_coverage(rows))
    write_csv(TABLE_DIR / "mrv_year_scope_counts.csv", build_year_scope_counts(rows))
    write_csv(TABLE_DIR / "mrv_processed_summary.csv", build_summary(rows))
    make_figures(rows)


if __name__ == "__main__":
    main()
