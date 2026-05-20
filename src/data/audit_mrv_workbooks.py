#!/usr/bin/env python3
"""Audit raw THETIS-MRV public emission report workbooks.

The public workbooks use two header rows: row 1 contains section/group labels
and row 3 contains field labels. This script flattens those headers, records
sheet inventory, and computes lightweight missingness/type summaries.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
RAW_DIR = ROOT / "data" / "raw" / "mrv"
TABLE_DIR = ROOT / "reports" / "tables"


def normalize(value: str) -> str:
    value = value.strip().lower()
    replacements = {
        "co₂": "co2",
        "co2": "co2",
        "co₂eq": "co2eq",
        "ch₄": "ch4",
        "n₂o": "n2o",
        "·": " ",
        "³": "3",
    }
    for old, new in replacements.items():
        value = value.replace(old, new)
    value = re.sub(r"\[[^\]]+\]", "", value)
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return value.strip("_")


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def stringify(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def load_metadata() -> dict[tuple[int, int], dict[str, Any]]:
    candidates = sorted(RAW_DIR.glob("downloadable-files-*.json"))
    if not candidates:
        return {}
    with candidates[-1].open(encoding="utf-8") as fh:
        data = json.load(fh)
    out = {}
    for item in data.get("results", []):
        out[(int(item["reportingPeriod"]), int(item["version"]))] = item
    return out


def version_from_name(path: Path) -> tuple[int | None, int | None]:
    m = re.match(r"(?P<year>\d{4})-v(?P<version>\d+)-", path.name)
    if not m:
        return None, None
    return int(m.group("year")), int(m.group("version"))


def sheet_scope(sheet_name: str) -> str:
    name = sheet_name.lower()
    if "partial" in name:
        return "partial_er"
    if "full" in name:
        return "full_er"
    return "annual_er"


def flattened_headers(ws) -> list[dict[str, Any]]:
    group_row = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    field_row = list(next(ws.iter_rows(min_row=3, max_row=3, values_only=True)))
    group = ""
    seen = Counter()
    headers = []
    for idx, (raw_group, raw_field) in enumerate(zip(group_row, field_row), start=1):
        if raw_group is not None:
            group = stringify(raw_group)
        if raw_field is None:
            continue
        field = stringify(raw_field)
        base_norm = f"{normalize(group)}__{normalize(field)}"
        seen[base_norm] += 1
        norm = base_norm if seen[base_norm] == 1 else f"{base_norm}_{seen[base_norm]}"
        headers.append(
            {
                "column_index": idx,
                "group": group,
                "raw_field": field,
                "normalized_field": norm,
            }
        )
    return headers


def audit() -> None:
    TABLE_DIR.mkdir(parents=True, exist_ok=True)
    metadata = load_metadata()

    inventory_rows = []
    schema_rows = []
    presence = defaultdict(set)
    ship_type_counts = Counter()

    for path in sorted(RAW_DIR.glob("*.xlsx")):
        year, version = version_from_name(path)
        file_hash = sha256(path)
        meta = metadata.get((year or -1, version or -1), {})
        workbook = load_workbook(path, read_only=True, data_only=True)

        for ws in workbook.worksheets:
            headers = flattened_headers(ws)
            non_missing = [0] * len(headers)
            examples = [""] * len(headers)
            type_counts = [Counter() for _ in headers]
            row_count = 0

            for row in ws.iter_rows(min_row=4, values_only=True):
                if not any(value is not None and value != "" for value in row):
                    continue
                row_count += 1
                for header in headers:
                    if header["normalized_field"] == "ship__ship_type":
                        idx = header["column_index"] - 1
                        ship_type_counts[stringify(row[idx])] += 1
                        break
                for j, header in enumerate(headers):
                    idx = header["column_index"] - 1
                    value = row[idx] if idx < len(row) else None
                    if value is None or value == "":
                        continue
                    non_missing[j] += 1
                    if not examples[j]:
                        examples[j] = stringify(value)
                    type_counts[j][type(value).__name__] += 1

            inventory_rows.append(
                {
                    "file": path.name,
                    "sheet": ws.title,
                    "reporting_period": year or "",
                    "version": version or "",
                    "scope": sheet_scope(ws.title),
                    "rows": row_count,
                    "columns": len(headers),
                    "generation_date": meta.get("generationDate", ""),
                    "sha256": file_hash,
                }
            )

            for j, header in enumerate(headers):
                missing_pct = 1 - (non_missing[j] / row_count) if row_count else 1
                schema_row = {
                    "file": path.name,
                    "sheet": ws.title,
                    "reporting_period": year or "",
                    "scope": sheet_scope(ws.title),
                    **header,
                    "non_missing": non_missing[j],
                    "missing_pct": f"{missing_pct:.4f}",
                    "type_counts": json.dumps(type_counts[j], ensure_ascii=False, sort_keys=True),
                    "sample_value": examples[j],
                }
                schema_rows.append(schema_row)
                presence[header["normalized_field"]].add(ws.title)

    write_csv(TABLE_DIR / "mrv_workbook_inventory.csv", inventory_rows)
    write_csv(TABLE_DIR / "mrv_schema_audit.csv", schema_rows)

    presence_rows = [
        {
            "normalized_field": field,
            "sheet_count": len(sheets),
            "sheets": "; ".join(sorted(sheets)),
        }
        for field, sheets in sorted(presence.items())
    ]
    write_csv(TABLE_DIR / "mrv_column_presence.csv", presence_rows)

    ship_type_rows = [
        {"ship_type": ship_type, "rows": count}
        for ship_type, count in ship_type_counts.most_common()
    ]
    write_csv(TABLE_DIR / "mrv_ship_type_counts.csv", ship_type_rows)


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


if __name__ == "__main__":
    audit()
